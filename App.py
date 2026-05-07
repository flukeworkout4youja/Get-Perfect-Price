import pandas as pd
from openpyxl import load_workbook
import os
from datetime import datetime
import re
import sys

# =========================================================
# BASE PATH (รองรับ .exe / .app / python)
# =========================================================

BASE_DIR = os.path.dirname(sys.executable) if getattr(sys, 'frozen', False) else os.path.dirname(os.path.abspath(__file__))

IN_FOLDER = os.path.join(BASE_DIR, "in")
OUT_FOLDER = os.path.join(BASE_DIR, "out")
TEMPLATE_FILE = os.path.join(BASE_DIR, "template.xlsx")

# =========================================================

def find_input_file():

    if not os.path.exists(IN_FOLDER):
        os.makedirs(IN_FOLDER)

    files = [f for f in os.listdir(IN_FOLDER) if f.endswith(".xlsx")]
    files = [f for f in files if "template" not in f.lower() and "~$" not in f]

    if not files:
        print("❌ ไม่มีไฟล์ใน in/")
        return None

    files.sort(
        key=lambda x: os.path.getmtime(os.path.join(IN_FOLDER, x)),
        reverse=True
    )

    return os.path.join(IN_FOLDER, files[0])

# =========================================================

def split_bill(text):

    if pd.isna(text):
        return ("", 0)

    text = str(text).strip()

    match = re.match(r'([A-Za-z]+)(\d+)', text)

    if match:
        return match.group(1), int(match.group(2))

    return text, 0

# =========================================================

def process():

    input_file = find_input_file()

    if not input_file:
        return

    print(f"📥 ใช้ไฟล์: {input_file}")

    # =====================================================
    # CHECK TEMPLATE
    # =====================================================

    if not os.path.exists(TEMPLATE_FILE):
        print("❌ ไม่พบ template.xlsx")
        return

    # =====================================================
    # READ EXCEL
    # =====================================================

    try:
        df = pd.read_excel(input_file, dtype=str)
    except Exception as e:
        print(f"❌ เปิดไฟล์ไม่ได้: {e}")
        return

    df.columns = df.columns.astype(str).str.strip()

    # =====================================================
    # FIND COLUMN
    # =====================================================

    def find_col(names):

        for n in names:
            for c in df.columns:
                if c.strip().lower() == n.strip().lower():
                    return c

        return None

    col = {
        'date': find_col(['Create date', 'Date', 'วันที่']),
        'bill': find_col(['Bill No.', 'Invoice No']),
        'product_code': find_col(['Product code', 'Product Code']),
        'qty': find_col(['Quantity', 'Qty']),
        'price': find_col(['Price/Unit', 'Price']),
        'discount': find_col(['Discount'])
    }

    # =====================================================
    # CHECK REQUIRED COLUMN
    # =====================================================

    for key in ['date', 'bill', 'product_code', 'qty', 'price']:

        if col[key] is None:
            print(f"❌ ขาด column: {key}")
            return

    # =====================================================
    # FILL DOWN
    # =====================================================

    df[col['date']] = df[col['date']].ffill()
    df[col['bill']] = df[col['bill']].ffill()

    # =====================================================
    # CLEAN STRING
    # =====================================================

    df[col['bill']] = df[col['bill']].astype(str).str.strip()
    df[col['product_code']] = df[col['product_code']].astype(str).str.strip()

    # =====================================================
    # FILTER PRODUCT
    # =====================================================

    df = df[
        df[col['product_code']].notna()
        & (df[col['product_code']] != '')
        & (df[col['product_code']].str.lower() != 'nan')
    ].reset_index(drop=True)

    if len(df) == 0:
        print("❌ ไม่มีข้อมูลสินค้า")
        return

    # =====================================================
    # CONVERT TYPE
    # =====================================================

    df['DATE'] = pd.to_datetime(
        df[col['date']],
        errors='coerce',
        dayfirst=True
    )

    df['DATE_STR'] = df['DATE'].dt.strftime('%Y%m%d').fillna('19000101')

    df['QTY'] = pd.to_numeric(
        df[col['qty']],
        errors='coerce'
    ).fillna(1)

    df['PRICE'] = pd.to_numeric(
        df[col['price']],
        errors='coerce'
    ).fillna(0)

    if col['discount']:

        df['DISC'] = pd.to_numeric(
            df[col['discount']],
            errors='coerce'
        ).fillna(0)

    else:
        df['DISC'] = 0

    # =====================================================
    # SORT ทีละวัน + PREFIX + NUMBER
    # =====================================================

    final_df = []

    for date, group in df.groupby('DATE', dropna=False):

        bill_df = group[[col['bill']]].drop_duplicates().copy()

        bill_df[['PREFIX', 'NUMBER']] = bill_df[col['bill']].apply(
            lambda x: pd.Series(split_bill(x))
        )

        bill_df = bill_df.sort_values(
            by=['PREFIX', 'NUMBER'],
            ascending=[True, True]
        ).reset_index(drop=True)

        ordered_bills = bill_df[col['bill']].tolist()

        for b in ordered_bills:

            sub = group[group[col['bill']] == b]

            final_df.append(sub)

    df = pd.concat(final_df).reset_index(drop=True)

    # =====================================================
    # GROUP ID
    # =====================================================

    df['GROUP_ID'] = (
        df[col['bill']] != df[col['bill']].shift()
    ).cumsum()

    # =====================================================
    # LOAD TEMPLATE
    # =====================================================

    try:
        wb = load_workbook(TEMPLATE_FILE)
    except Exception as e:
        print(f"❌ เปิด template ไม่ได้: {e}")
        return

    ws = wb.active

    start_row = 2

    # =====================================================
    # WRITE EXCEL
    # =====================================================

    for i, row in df.iterrows():

        excel_row = start_row + i

        ws.cell(
            row=excel_row,
            column=1,
            value=int(row['GROUP_ID'])
        )

        is_first = (
            i == 0
            or row[col['bill']] != df.loc[i - 1, col['bill']]
        )

        # =================================================
        # FIRST ROW OF BILL
        # =================================================

        if is_first:

            ws.cell(
                row=excel_row,
                column=2,
                value=row['DATE_STR']
            )

            ws.cell(
                row=excel_row,
                column=3,
                value=row[col['bill']]
            )

            ws.cell(
                row=excel_row,
                column=5,
                value='V00001'
            )

            ws.cell(
                row=excel_row,
                column=8,
                value=2
            )

            ws.cell(
                row=excel_row,
                column=9,
                value=2
            )

            ws.cell(
                row=excel_row,
                column=18,
                value='CSH001'
            )

        # =================================================
        # EVERY ROW
        # =================================================

        ws.cell(
            row=excel_row,
            column=10,
            value=str(row[col['product_code']])
        )

        ws.cell(
            row=excel_row,
            column=13,
            value=float(row['QTY'])
        )

        ws.cell(
            row=excel_row,
            column=14,
            value=float(row['PRICE'])
        )

        ws.cell(
            row=excel_row,
            column=15,
            value=float(row['DISC'])
        )

        ws.cell(
            row=excel_row,
            column=16,
            value='7%'
        )

    # =====================================================
    # SAVE FILE
    # =====================================================

    if not os.path.exists(OUT_FOLDER):
        os.makedirs(OUT_FOLDER)

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")

    out_file = os.path.join(
        OUT_FOLDER,
        f"result_{ts}.xlsx"
    )

    try:
        wb.save(out_file)
    except PermissionError:
        print("❌ บันทึกไม่ได้ กรุณาปิดไฟล์ Excel ก่อน")
        return

    print(f"✅ เสร็จ: {out_file}")

# =========================================================

if __name__ == "__main__":

    process()

    input("\nกด Enter เพื่อปิด...")
