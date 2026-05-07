import pandas as pd
from openpyxl import load_workbook
import os
from datetime import datetime
import re

IN_FOLDER = "in"
OUT_FOLDER = "out"
TEMPLATE_FILE = "template.xlsx"

def find_input_file():
    files = [f for f in os.listdir(IN_FOLDER) if f.endswith(".xlsx")]
    files = [f for f in files if "template" not in f.lower() and "~$" not in f]

    if not files:
        print("❌ ไม่มีไฟล์ใน in/")
        return None

    files.sort(key=lambda x: os.path.getmtime(os.path.join(IN_FOLDER, x)), reverse=True)
    return os.path.join(IN_FOLDER, files[0])

def split_bill(text):
    if pd.isna(text):
        return ("", 0)
    text = str(text)
    match = re.match(r'([A-Za-z]+)(\d+)', text)
    if match:
        return match.group(1), int(match.group(2))
    return text, 0

def process():

    input_file = find_input_file()
    if not input_file:
        return

    print(f"📥 ใช้ไฟล์: {input_file}")

    df = pd.read_excel(input_file)
    df.columns = df.columns.str.strip()

    def find_col(names):
        for n in names:
            if n in df.columns:
                return n
        return None

    col = {
        'date': find_col(['Create date','Date','วันที่']),
        'bill': find_col(['Bill No.','Invoice No']),
        'product_code': find_col(['Product code','Product Code']),
        'qty': find_col(['Quantity','Qty']),
        'price': find_col(['Price/Unit','Price']),
        'discount': find_col(['Discount'])
    }

    for key in ['date','bill','product_code','qty','price']:
        if col[key] is None:
            print(f"❌ ขาด column: {key}")
            return

    # fill down
    df[col['date']] = df[col['date']].ffill()
    df[col['bill']] = df[col['bill']].ffill()

    # filter
    df = df[df[col['product_code']].notna()].reset_index(drop=True)

    # convert
    df['DATE'] = pd.to_datetime(df[col['date']], errors='coerce', dayfirst=True)
    df['DATE_STR'] = df['DATE'].dt.strftime('%Y%m%d').fillna('19000101')

    df['QTY'] = pd.to_numeric(df[col['qty']], errors='coerce').fillna(1)
    df['PRICE'] = pd.to_numeric(df[col['price']], errors='coerce').fillna(0)

    if col['discount']:
        df['DISC'] = pd.to_numeric(df[col['discount']], errors='coerce').fillna(0)
    else:
        df['DISC'] = 0

    # =========================================================
    # SORT ทีละวัน + เรียง bill
    # =========================================================

    final_df = []

    for date, group in df.groupby('DATE'):

        bill_df = group[[col['bill']]].drop_duplicates().copy()

        bill_df[['PREFIX', 'NUMBER']] = bill_df[col['bill']].apply(
            lambda x: pd.Series(split_bill(x))
        )

        bill_df = bill_df.sort_values(
            by=['PREFIX', 'NUMBER'],
            ascending=[True, True]
        ).reset_index(drop=True)

        ordered_bills = bill_df[col['bill']].tolist()

        for b in ordered_bills:
            sub = group[group[col['bill']] == b]
            final_df.append(sub)

    df = pd.concat(final_df).reset_index(drop=True)

    # group id
    df['GROUP_ID'] = (df[col['bill']] != df[col['bill']].shift()).cumsum()

    # -------------------------------
    # write excel
    # -------------------------------
    wb = load_workbook(TEMPLATE_FILE)
    ws = wb.active

    start_row = 2

    for i, row in df.iterrows():

        excel_row = start_row + i

        ws.cell(row=excel_row, column=1, value=row['GROUP_ID'])

        is_first = (i == 0) or (row[col['bill']] != df.loc[i-1, col['bill']])

        if is_first:
            ws.cell(row=excel_row, column=2, value=row['DATE_STR'])
            ws.cell(row=excel_row, column=3, value=row[col['bill']])
            ws.cell(row=excel_row, column=5, value='V00001')
            ws.cell(row=excel_row, column=8, value=2)
            ws.cell(row=excel_row, column=9, value=2)
            ws.cell(row=excel_row, column=18, value='CSH001')  # 🔥 เพิ่มตรงนี้ (คอลัม R)

        ws.cell(row=excel_row, column=10, value=row[col['product_code']])
        ws.cell(row=excel_row, column=13, value=row['QTY'])
        ws.cell(row=excel_row, column=14, value=row['PRICE'])
        ws.cell(row=excel_row, column=15, value=row['DISC'])
        ws.cell(row=excel_row, column=16, value='7%')

    os.makedirs(OUT_FOLDER, exist_ok=True)

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_file = os.path.join(OUT_FOLDER, f"result_{ts}.xlsx")

    wb.save(out_file)

    print(f"✅ เสร็จ: {out_file}")


if __name__ == "__main__":
    process()
    input("\nกด Enter เพื่อปิด...")